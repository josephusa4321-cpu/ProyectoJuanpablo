import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def blindar_excel():
    excel_path = "Estado_de_Cuenta_Flota.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    
    # 1. Crear Hoja de Instrucciones
    if "INSTRUCCIONES" in wb.sheetnames:
        del wb["INSTRUCCIONES"]
    
    ws = wb.create_sheet("INSTRUCCIONES", 0)
    ws.sheet_properties.tabColor = "38BDF8"
    
    # Diseño de la cabecera
    ws['B2'] = "GUÍA RÁPIDA: GENERACIÓN DE REPORTES"
    ws['B2'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['B2'].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    ws.merge_cells('B2:F3')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # Pasos
    pasos = [
        ("PASO 1", "Actualizar Movimientos", "Llene sus ingresos y gastos en la hoja 'Movimientos'. Use los desplegables."),
        ("PASO 2", "Guardar y Cerrar", "Guarde los cambios (Ctrl + G) y cierre el archivo Excel."),
        ("PASO 3", "Doble Clic", "Haga doble clic en el archivo 'GENERAR_REPORTES.bat' en su carpeta."),
        ("PASO 4", "¡Listo!", "Sus reportes estarán en la carpeta 'reportes_generados' listos para enviar.")
    ]

    for i, (title, action, desc) in enumerate(pasos):
        row = 5 + (i * 3)
        ws.cell(row=row, column=2, value=title).font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=action).font = Font(bold=True, size=12, color="38BDF8")
        ws.cell(row=row+1, column=3, value=desc).font = Font(size=10)

    # 4. Ajustar anchos
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60

    wb.save(excel_path)
    print("Hoja de instrucciones inyectada con éxito.")

if __name__ == "__main__":
    blindar_excel()
